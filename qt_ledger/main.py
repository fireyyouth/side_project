import sys
import sqlite3
from typing import override
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QTabWidget,
    QTableWidget, QTableWidgetItem, QHBoxLayout, QLineEdit, QFormLayout,
    QDialog, QDialogButtonBox, QLabel, QMessageBox, QDateTimeEdit, QComboBox, 
    QHeaderView, QFrame, QTreeWidget, QTreeWidgetItem, QSizePolicy, QSpacerItem,
    QFileDialog
)
from PySide6.QtGui import QDoubleValidator, QFont, QDropEvent, QDragMoveEvent, QDragEnterEvent, QDragLeaveEvent, QDrag
from PySide6.QtCore import Qt
from PySide6.QtCore import QDateTime
from collections import defaultdict
from functools import partial
import decimal
import sys
import openpyxl

conn = sqlite3.connect("ledger.db")
cursor = conn.cursor()

def init_db(drop):
    cursor.execute("PRAGMA foreign_keys = ON")

    if drop:
        cursor.execute("DROP TABLE IF EXISTS person")
        cursor.execute("DROP TABLE IF EXISTS project")
        cursor.execute("DROP TABLE IF EXISTS sub_project")
        cursor.execute("DROP TABLE IF EXISTS transfer")

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS person (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            UNIQUE(name)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS project (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            rank INTEGER NOT NULL,
            UNIQUE(name)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sub_project (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            parent INTEGER NOT NULL,
            rank INTEGER NOT NULL,
            UNIQUE(name, parent),
            FOREIGN KEY (parent) REFERENCES project(id) ON DELETE RESTRICT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS transfer (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            time TEXT NOT NULL,
            person INTEGER NOT NULL,
            sub_project INTEGER NOT NULL,
            kind TEXT NOT NULL,
            amount TEXT NOT NULL,
            memo TEXT,
            FOREIGN KEY (person) REFERENCES person(id) ON DELETE RESTRICT,
            FOREIGN KEY (sub_project) REFERENCES sub_project(id) ON DELETE RESTRICT
        )
    """)
    conn.commit()

def get_person():
    cursor.execute("SELECT id, name FROM person")
    return cursor.fetchall()

def get_person_name(id):
    cursor.execute("SELECT name FROM person WHERE id = ?", (id,))
    return cursor.fetchone()[0]

def get_project():
    cursor.execute("SELECT id, name FROM project ORDER BY rank")
    return cursor.fetchall()

def swap_project_order(name1, name2):
    cursor.execute("SELECT rank FROM project WHERE name = ?", (name1,))
    rank1 = cursor.fetchone()[0]
    cursor.execute("SELECT rank FROM project WHERE name = ?", (name2,))
    rank2 = cursor.fetchone()[0]

    cursor.execute("BEGIN")
    cursor.execute("UPDATE project SET rank = ? WHERE name = ?", (rank2, name1))
    cursor.execute("UPDATE project SET rank = ? WHERE name = ?", (rank1, name2))
    conn.commit()

def get_sub_project(parent=None):
    if parent != None:
        print(parent)
        cursor.execute("SELECT id FROM project WHERE name = ?", (parent,))
        item = cursor.fetchone()
        if item is None:
            return []
        parent_id = item[0]
        cursor.execute("""
            SELECT sub_project.name AS name, project.name AS parent
            FROM sub_project LEFT JOIN project ON sub_project.parent = project.id
            WHERE parent = ?
            ORDER BY project.rank ASC, sub_project.rank ASC
        """, (parent_id,))
    else:
        cursor.execute("""
            SELECT sub_project.name AS name, project.name AS parent
            FROM sub_project LEFT JOIN project ON sub_project.parent = project.id
            ORDER BY project.rank ASC, sub_project.rank ASC
        """)
    return cursor.fetchall()


def swap_sub_project_order(parent, name1, name2):
    cursor.execute("SELECT rank FROM sub_project WHERE name = ? AND parent = ?", (name1, parent))
    rank1 = cursor.fetchone()[0]
    cursor.execute("SELECT rank FROM sub_project WHERE name = ? AND parent = ?", (name2, parent))
    rank2 = cursor.fetchone()[0]

    cursor.execute("BEGIN")
    cursor.execute("UPDATE sub_project SET rank = ? WHERE name = ? AND parent = ?", (rank2, name1, parent))
    cursor.execute("UPDATE sub_project SET rank = ? WHERE name = ? AND parent = ?", (rank1, name2, parent))
    conn.commit()

def add_person(name):
    if name:
        cursor.execute("INSERT INTO person (name) VALUES (?)", (name,))
        conn.commit()

def update_person(person_id, name):
    if name:
        cursor.execute("UPDATE person SET name = ? WHERE id = ?", (name, person_id))
        conn.commit()

def add_project(name):
    if name:
        cursor.execute("SELECT MAX(rank) FROM project")
        item = cursor.fetchone()[0]
        print(item)
        max_rank = item if item else 0
        cursor.execute("INSERT INTO project (name, rank) VALUES (?, ?)", (name, max_rank + 1))
        conn.commit()

def update_project(name, new_name):
    if new_name:
        cursor.execute("UPDATE project SET name = ? WHERE name = ?", (new_name, name))
        conn.commit()

def add_sub_project(name, parent):
    if name:
        cursor.execute("SELECT MAX(rank) FROM sub_project WHERE parent = ?", (parent,))
        item = cursor.fetchone()[0]
        max_rank = item if item else 0
        cursor.execute("INSERT INTO sub_project (name, parent, rank) VALUES (?, ?, ?)", (name, parent, max_rank + 1))
        conn.commit()

def update_sub_project(parent, name, new_name):
    if new_name:
        cursor.execute("UPDATE sub_project SET name = ? WHERE parent = ? AND name = ?", (new_name, parent, name))
        conn.commit()

def delete_person(person):
    cursor.execute("DELETE FROM person WHERE name=?", (person,))
    conn.commit()

def delete_project(project):
    cursor.execute("DELETE FROM project WHERE name=?", (project,))
    conn.commit()


def delete_sub_project(project, sub_project):
    cursor.execute("DELETE FROM sub_project WHERE parent=? AND name=?", (project, sub_project))
    conn.commit()

class InvalidInputError(Exception):
    pass

class BalanceError(Exception):
    pass

def kind_sign(kind):
    return 1 if kind == '入账' else -1

def get_balance(person, project):
    cursor.execute("""
        SELECT sub_project.name, transfer.amount, transfer.kind
        FROM transfer
        LEFT JOIN person ON transfer.person = person.id
        LEFT JOIN sub_project ON transfer.sub_project = sub_project.id
        LEFT JOIN project ON sub_project.parent = project.id
        WHERE person.name=? AND project.name=?
    """, (person, project))
    sub_project_balance = defaultdict(decimal.Decimal)
    for (sub_project, amount, kind) in cursor.fetchall():
        sub_project_balance[sub_project] += kind_sign(kind) * decimal.Decimal(amount)
    return sub_project_balance

def post_check_balance(person, project, sub_project):
    balance = get_balance(person, project).get(sub_project, 0)
    print('post check balance', balance)
    if balance < 0:
        raise BalanceError(f'{person} 在 {project} {sub_project} 上的余额会变成 {balance}')

def person_name_to_id(name):
    item = cursor.execute("SELECT id FROM person WHERE name=?", (name,)).fetchone()
    if item is None:
        raise InvalidInputError(f'人员 {name} 不存在')
    return item[0]

def project_name_to_id(parent_name, name):
    item = cursor.execute("SELECT id FROM project WHERE name=?", (parent_name,)).fetchone()
    if item is None:
        raise InvalidInputError(f'项目 {parent_name} 不存在')
    project_id = item[0]

    item = cursor.execute("SELECT id FROM sub_project WHERE name=? AND parent=?", (name, project_id)).fetchone()
    if item is None:
        raise InvalidInputError(f'子项目 {name} 不存在')
    sub_project_id = item[0]

    return sub_project_id

def add_transfer(time, person, project, sub_project, kind, amount, memo):
    person_id = person_name_to_id(person)
    sub_project_id = project_name_to_id(project, sub_project)

    try:
        cursor.execute('BEGIN')
        cursor.execute(
            "INSERT INTO transfer (time, person, sub_project, kind, amount, memo) VALUES (?, ?, ?, ?, ?, ?)",
            (time, person_id, sub_project_id, kind, amount, memo))
        post_check_balance(person, project, sub_project)
        conn.commit()
    except Exception:
        conn.rollback()
        raise

def delete_transfer(id_):
    person, project, sub_project = cursor.execute("""
        SELECT person.name, project.name, sub_project.name
        FROM transfer
        LEFT JOIN person ON transfer.person = person.id
        LEFT JOIN sub_project ON transfer.sub_project = sub_project.id
        LEFT JOIN project ON sub_project.parent = project.id
        WHERE id = ?
    """, (id_,)).fetchall()[0]

    try:
        cursor.execute('BEGIN')
        cursor.execute("DELETE FROM transfer WHERE id=?", (id_,))
        post_check_balance(person, project, sub_project)
        conn.commit()
    except Exception:
        conn.rollback()
        raise

def update_transfer(id_, time, person, project, sub_project, kind, amount, memo):
    person_id = person_name_to_id(person)
    sub_project_id = project_name_to_id(project, sub_project)

    old_person, old_project, old_sub_project = cursor.execute("""
        SELECT person.name, project.name, sub_project.name
        FROM transfer
        LEFT JOIN person ON transfer.person = person.id
        LEFT JOIN sub_project ON transfer.sub_project = sub_project.id
        LEFT JOIN project ON sub_project.parent = project.id
        WHERE id = ?
    """, (id_,)).fetchall()[0]

    try:
        cursor.execute('BEGIN')
        cursor.execute("""
            UPDATE transfer
            SET time = ?, person = ?, sub_project = ?, kind = ?, amount = ?, memo = ?
            WHERE id = ?
        """, (time, person_id, sub_project_id, kind, amount, memo, id_))
        for p in [old_person, person]:
            for j, k in [(old_project, old_sub_project), (project, sub_project)]:
                post_check_balance(p, j, k)

        conn.commit()
    except Exception:
        conn.rollback()
        raise

def get_transfer():
    stmt = """
        SELECT transfer.id, transfer.time, person.name, project.name, sub_project.name, transfer.kind, transfer.amount, transfer.memo
        FROM transfer
        LEFT JOIN person ON transfer.person = person.id
        LEFT JOIN sub_project ON transfer.sub_project = sub_project.id
        LEFT JOIN project ON sub_project.parent = project.id
        ORDER BY time DESC
    """
    print('sql', stmt)
    cursor.execute(stmt)
    return cursor.fetchall()

def filter_transfer(person, project, kind):
    stmt = '''
        SELECT transfer.id, transfer.time, person.name, project.name, sub_project.name, transfer.kind, transfer.amount, transfer.memo
        FROM transfer
        LEFT JOIN person ON transfer.person = person.id
        LEFT JOIN sub_project ON transfer.sub_project = sub_project.id
        LEFT JOIN project ON sub_project.parent = project.id
        WHERE 1 = 1
    '''
    if person:
        stmt += f' AND person = "{person}"'
    if project:
        stmt += f' AND project = "{project}"'
    if kind:
        stmt += f' AND kind = "{kind}"'
    stmt += ' ORDER BY time DESC'
    print('sql', stmt)
    cursor.execute(stmt)
    return cursor.fetchall()


def create_person_combo():
    combo = QComboBox()
    for _, name in get_person():
        combo.addItem(name)
    return combo

def create_project_combo():
    combo = QComboBox()
    for _, name in get_project():
        combo.addItem(name)
    return combo

def create_kind_combo(need_blank=False):
    combo = QComboBox()
    if need_blank:
        combo.addItem('')
    combo.addItems(["入账", "出账"])
    return combo


def excel_from_table(table: QTableWidget, title, export_vertical_header):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # Get the headers from the table and add them to the first row in the sheet
    for col in range(table.columnCount()):
        ws.cell(row=1, column=col+2, value=table.horizontalHeaderItem(col).text())

    # Get the vertical headers (row headers) and add them to the first column
    if export_vertical_header:
        for row in range(table.rowCount()):
            ws.cell(row=row+2, column=1, value=table.verticalHeaderItem(row).text())

    # Get the table data and add it to the sheet starting from row 2
    for row in range(table.rowCount()):
        for col in range(table.columnCount()):
            item = table.item(row, col)
            if item:
                ws.cell(row=row+2, column=col+2, value=item.text())

    return wb

class EditTranferDialog(QDialog):
    def __init__(self, id_, time, person, project, sub_project, kind, amount, memo):
        super().__init__()

        self.id_ = id_

        self.setWindowTitle("编辑")

        self.time_edit = QDateTimeEdit()
        self.amount_edit = QLineEdit()
        self.amount_edit.setValidator(QDoubleValidator(0.0, float('inf'), 2))
        self.person_combo = create_person_combo()
        self.project_combo = create_project_combo()
        self.sub_project_combo = QComboBox()
        self.kind_combo = create_kind_combo()
        self.memo_edit = QLineEdit()

        self.load_sub_projects()
        self.person_combo.setEditable(True) # allow deleted person
        self.project_combo.setEditable(True) # allow deleted project
        self.sub_project_combo.setEditable(True) # allow deleted sub project

        self.time_edit.setDateTime(QDateTime.fromString(time, "yyyy-MM-dd HH:mm:ss"))
        self.amount_edit.setText(str(amount))
        self.person_combo.setCurrentText(person)
        self.project_combo.setCurrentText(project)
        self.sub_project_combo.setCurrentText(sub_project)
        self.kind_combo.setCurrentText(kind)
        self.memo_edit.setText(memo)
        
        self.project_combo.currentTextChanged.connect(self.load_sub_projects)

        form_layout = QVBoxLayout()
        form_layout.addWidget(QLabel("时间:"))
        form_layout.addWidget(self.time_edit)
        form_layout.addWidget(QLabel("人员:"))
        form_layout.addWidget(self.person_combo)
        form_layout.addWidget(QLabel("项目:"))
        form_layout.addWidget(self.project_combo)
        form_layout.addWidget(QLabel("子项目:"))
        form_layout.addWidget(self.sub_project_combo)
        form_layout.addWidget(QLabel("类型:"))
        form_layout.addWidget(self.kind_combo)
        form_layout.addWidget(QLabel("金额:"))
        form_layout.addWidget(self.amount_edit)
        form_layout.addWidget(QLabel("备注:"))
        form_layout.addWidget(self.memo_edit)

        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.handle_save)
        button_box.rejected.connect(self.reject)

        layout = QVBoxLayout()
        layout.addLayout(form_layout)
        layout.addWidget(button_box)
        self.setLayout(layout)

    def load_sub_projects(self):
        self.sub_project_combo.clear()
        for name, _ in get_sub_project(self.project_combo.currentText()):
            self.sub_project_combo.addItem(name)

    def handle_save(self):
        time = self.time_edit.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        person = self.person_combo.currentText()
        project = self.project_combo.currentText()
        sub_project = self.sub_project_combo.currentText()
        kind = self.kind_combo.currentText()
        amount = self.amount_edit.text()
        memo = self.memo_edit.text()
        try:
            update_transfer(self.id_, time, person, project, sub_project, kind, amount, memo)
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))
            return
        self.accept()

class EditPersonDialog(QDialog):
    def __init__(self, person_id):
        super().__init__()
        self.person_id = person_id
        self.setWindowTitle("编辑姓名")
        self.name_input = QLineEdit()
        self.name_input.setText(get_person_name(person_id))
        self.save_btn = QPushButton("保存")
        self.save_btn.clicked.connect(self.handle_save)
        layout = QVBoxLayout()
        layout.addWidget(self.name_input)
        layout.addWidget(self.save_btn)
        self.setLayout(layout)

    def handle_save(self):
        name = self.name_input.text()
        if name:
            update_person(self.person_id, name)
            self.accept()

class PersonTab(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()
        form = QFormLayout()
        self.person_name_input = QLineEdit()
        form.addRow("姓名:", self.person_name_input)
        add_btn = QPushButton("添加")
        add_btn.clicked.connect(self.handle_add)
        layout.addLayout(form)
        layout.addWidget(add_btn)
        self.person_table = QTableWidget()
        self.person_table.setColumnCount(2)
        self.person_table.setHorizontalHeaderLabels(["姓名", "操作"])
        self.person_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.person_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.person_table)
        self.setLayout(layout)
        self.load()

    def load(self):
        rows = get_person()
        self.person_table.setRowCount(len(rows))
        for row, (id_, name) in enumerate(rows):
            self.person_table.setItem(row, 0, QTableWidgetItem(name))

            delete_btn = QPushButton("删除")
            delete_btn.clicked.connect(partial(self.handle_delete, name))

            edit_btn = QPushButton("编辑")
            edit_btn.clicked.connect(partial(self.handle_edit, id_))

            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(0, 0, 0, 0)
            action_layout.setSpacing(2)

            action_widget = QWidget()
            action_widget.setLayout(action_layout)
            action_layout.addWidget(edit_btn)
            action_layout.addWidget(delete_btn)

            self.person_table.setCellWidget(row, 1, action_widget)

    def handle_delete(self, name):
        reply = QMessageBox.question(self, "确认", f"确定要删除 {name} 吗？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return

        try:
            delete_person(name)
            self.load()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "错误", "存在相关交易记录，无法删除")
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))

    def handle_edit(self, person_id):
        dialog = EditPersonDialog(person_id)
        if dialog.exec():
            self.load()

    def handle_add(self):
        name = self.person_name_input.text()
        if name:
            add_person(name)
            self.load()

class CreateSubProjectDialog(QDialog):
    def __init__(self, parent_id):
        super().__init__()
        self.parent_id = parent_id
        self.setWindowTitle("新建子项目")
        self.name_input = QLineEdit()
        self.save_btn = QPushButton("保存")
        self.save_btn.clicked.connect(self.handle_save)
        layout = QVBoxLayout()
        layout.addWidget(self.name_input)
        layout.addWidget(self.save_btn)
        self.setLayout(layout)

    def handle_save(self):
        name = self.name_input.text()
        if name:
            add_sub_project(name, self.parent_id)
            self.accept()


class EditProjectDialog(QDialog):
    def __init__(self, name):
        super().__init__()
        self.name = name
        self.setWindowTitle("编辑项目")
        self.name_input = QLineEdit(self.name)
        self.save_btn = QPushButton("保存")
        self.save_btn.clicked.connect(self.handle_save)
        layout = QVBoxLayout()
        layout.addWidget(self.name_input)
        layout.addWidget(self.save_btn)
        self.setLayout(layout)

    def handle_save(self):
        name = self.name_input.text()
        if name:
            update_project(self.name, name)
            self.accept()

class EditSubProjectDialog(QDialog):
    def __init__(self, parent_id, name):
        super().__init__()
        self.parent_id = parent_id
        self.name = name
        self.setWindowTitle("编辑子项目")
        self.name_input = QLineEdit(self.name)
        self.save_btn = QPushButton("保存")
        self.save_btn.clicked.connect(self.handle_save)
        layout = QVBoxLayout()
        layout.addWidget(self.name_input)
        layout.addWidget(self.save_btn)
        self.setLayout(layout)

    def handle_save(self):
        name = self.name_input.text()
        if name:
            update_sub_project(self.parent_id, self.name, name)
            self.accept()

class ProjectTab(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()
        form = QFormLayout()
        self.project_name_input = QLineEdit()
        form.addRow("项目名:", self.project_name_input)
        add_btn = QPushButton("添加")

        add_btn.clicked.connect(self.handle_add)
        layout.addLayout(form)
        layout.addWidget(add_btn)
        self.project_tree = QTreeWidget()
        self.project_tree.setColumnCount(2)
        self.project_tree.setHeaderLabels(["项目名", "操作"])

        layout.addWidget(self.project_tree)
        self.setLayout(layout)
        self.load()

    def load(self):
        self.project_tree.clear()

        project_list = []
        project_id_map = {}
        for id, name in get_project():
            project_list.append(name)
            project_id_map[name] = id

        top_items = {}
        for i, name in enumerate(project_list):
            item = QTreeWidgetItem([name, ''])
            self.project_tree.addTopLevelItem(item)

            up_btn = QPushButton("上移")
            if i > 0:
                up_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
                up_btn.clicked.connect(partial(self.handle_move_project, name, project_list[i - 1]))
            else:
                up_btn.setDisabled(True)

            down_btn = QPushButton("下移")
            if i < len(project_list) - 1:
                down_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
                down_btn.clicked.connect(partial(self.handle_move_project, name, project_list[i + 1]))
            else:
                down_btn.setDisabled(True)

            create_btn = QPushButton("创建子项目")
            create_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            create_btn.clicked.connect(partial(self.handle_create_sub_project, project_id_map[name]))

            edit_btn = QPushButton("编辑")
            edit_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            edit_btn.clicked.connect(partial(self.handle_edit_project, name))

            delete_btn = QPushButton("删除")
            delete_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            delete_btn.clicked.connect(partial(self.handle_delete_project, name))

            action_widget = QWidget()
            action_layout = QHBoxLayout()
            if up_btn:
                action_layout.addWidget(up_btn)
            if down_btn:
                action_layout.addWidget(down_btn)
            action_layout.addWidget(edit_btn)
            action_layout.addWidget(delete_btn)
            action_layout.addWidget(create_btn)
            action_layout.addSpacerItem(QSpacerItem(0, 0, QSizePolicy.Expanding, QSizePolicy.Minimum))
            action_layout.setAlignment(edit_btn, Qt.AlignLeft)
            action_layout.setAlignment(delete_btn, Qt.AlignLeft)
            action_layout.setAlignment(create_btn, Qt.AlignLeft)
            action_widget.setLayout(action_layout)
            self.project_tree.setItemWidget(item, 1, action_widget)
            top_items[name] = item

        sub_project_map = defaultdict(list)
        for name, parent in get_sub_project():
            sub_project_map[parent].append(name)

        for parent, sub_project_list in sub_project_map.items():
            for i, name in enumerate(sub_project_list):
                item = QTreeWidgetItem([name, ''])
                top_items[parent].addChild(item)

                up_btn = QPushButton("上移")
                if i > 0:
                    up_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
                    up_btn.clicked.connect(partial(self.handle_move_sub_project, project_id_map[parent], name, sub_project_list[i - 1]))
                else:
                    up_btn.setDisabled(True)

                down_btn = QPushButton("下移")
                if i < len(sub_project_list) - 1:
                    down_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
                    down_btn.clicked.connect(partial(self.handle_move_sub_project, project_id_map[parent], name, sub_project_list[i + 1]))
                else:
                    down_btn.setDisabled(True)

                edit_btn = QPushButton("编辑")
                edit_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
                edit_btn.clicked.connect(partial(self.handle_edit_sub_project, project_id_map[parent], name))

                delete_btn = QPushButton("删除")
                delete_btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
                delete_btn.clicked.connect(partial(self.handle_delete_sub_project, project_id_map[parent], name))

                action_widget = QWidget()
                action_layout = QHBoxLayout()
                action_layout.addWidget(up_btn)
                action_layout.addWidget(down_btn)
                action_layout.addWidget(edit_btn)
                action_layout.addWidget(delete_btn)
                action_layout.addSpacerItem(QSpacerItem(0, 0, QSizePolicy.Expanding, QSizePolicy.Minimum))
                action_layout.setAlignment(up_btn, Qt.AlignLeft)
                action_layout.setAlignment(down_btn, Qt.AlignLeft)
                action_layout.setAlignment(edit_btn, Qt.AlignLeft)
                action_layout.setAlignment(delete_btn, Qt.AlignLeft)
                action_widget.setLayout(action_layout)
                self.project_tree.setItemWidget(item, 1, action_widget)

        self.project_tree.expandAll()

    def handle_move_project(self, name1, name2):
        swap_project_order(name1, name2)
        self.load()

    def handle_move_sub_project(self, parent, name1, name2):
        swap_sub_project_order(parent, name1, name2)
        self.load()

    def handle_create_sub_project(self, parent):
        dialog = CreateSubProjectDialog(parent)
        dialog.exec()
        self.load()

    def handle_add(self):
        name = self.project_name_input.text()
        if name:
            add_project(name)
            self.load()

    def handle_edit_project(self, name):
        dialog = EditProjectDialog(name)
        if dialog.exec() == QDialog.Accepted:
            self.load()

    def handle_add_sub(self):
        name = self.project_name_input.text()
        if name:
            add_project(name)
            self.load()

    def handle_edit_sub_project(self, parent_id, name):
        dialog = EditSubProjectDialog(parent_id, name)
        if dialog.exec() == QDialog.Accepted:
            self.load()

    def handle_delete_project(self, name):
        reply = QMessageBox.question(self, "确认", f"确定要删除 {name} 吗？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return

        try:
            delete_project(name)
            self.load()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "错误", "存在相关子项目，无法删除")
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))

    def handle_delete_sub_project(self, parent, name):
        reply = QMessageBox.question(self, "确认", f"确定要删除 {name} 吗？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return

        try:
            delete_sub_project(parent, name)
            self.load()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "错误", "存在相关交易记录，无法删除")
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))

class TransferTab(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()
        form = QFormLayout()

        self.time_input = QDateTimeEdit(QDateTime.currentDateTime())
        self.time_input.setCalendarPopup(True)
        self.person_input = QComboBox()
        self.project_input = QComboBox()
        self.project_balance = QLabel()
        self.sub_project_input = QComboBox()
        self.kind_input = create_kind_combo()
        self.amount_input = QLineEdit()
        self.amount_input.setValidator(QDoubleValidator(0.0, float('inf'), 2))
        self.memo_input = QLineEdit()

        self.person_input.setEditable(True)

        self.project_input.currentTextChanged.connect(self.load_sub_projects)
        self.project_input.currentTextChanged.connect(self.load_balance)
        self.person_input.currentTextChanged.connect(self.load_balance)

        form.addRow("时间:", self.time_input)
        form.addRow("人员:", self.person_input)
        form.addRow("项目:", self.project_input)
        form.addRow("余额:", self.project_balance)
        form.addRow("子项目:", self.sub_project_input)
        form.addRow("类型:", self.kind_input)
        form.addRow("金额:", self.amount_input)
        form.addRow("备注:", self.memo_input)

        add_btn = QPushButton("添加转账")
        add_btn.clicked.connect(self.handle_add)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)

        filter_bar = QHBoxLayout()
        self.person_filter = QLineEdit()
        self.project_filter = QLineEdit()
        self.kind_filter = create_kind_combo(True)

        filter_btn = QPushButton("过滤")

        export_btn = QPushButton('导出为 excel')
        export_btn.clicked.connect(self.export_to_excel)

        self.filters = ("", "", "")
        self.filters_label = QLabel()
        def handle_filter():
            self.filters = (self.person_filter.text(), self.project_filter.text(), self.kind_filter.currentText())
            self.load()
        filter_btn.clicked.connect(handle_filter)

        filter_bar.addWidget(QLabel("人员:"))
        filter_bar.addWidget(self.person_filter)
        filter_bar.addWidget(QLabel("项目:"))
        filter_bar.addWidget(self.project_filter)
        filter_bar.addWidget(QLabel("类型:"))
        filter_bar.addWidget(self.kind_filter)
        filter_bar.addWidget(filter_btn)

        layout.addLayout(form)
        layout.addWidget(add_btn)
        layout.addWidget(line)
        layout.addLayout(filter_bar)
        layout.addWidget(self.filters_label)
        layout.addWidget(export_btn)

        self.transfer_table = QTableWidget()
        headers = ["时间", "人员", "项目", "子项目", "类型", "金额", "备注", "操作"]
        self.transfer_table.setColumnCount(len(headers))
        self.transfer_table.setHorizontalHeaderLabels(headers)
    
        layout.addWidget(self.transfer_table)
        self.setLayout(layout)
        self.load()

    def load_balance(self):
        sub_project_balance = get_balance(self.person_input.currentText(), self.project_input.currentText())
        self.project_balance.clear()
        content = [f'合计: {sum(sub_project_balance.values())}']
        for sub_project, balance in sub_project_balance.items():
            content.append(f'{sub_project}: {balance}')
        self.project_balance.setText(', '.join(content))

    def load_sub_projects(self):
        self.sub_project_input.clear()
        for name, _ in get_sub_project(self.project_input.currentText()):
            self.sub_project_input.addItem(name)

    def load_combo(self):
        self.person_input.clear()
        for _, name in get_person():
            self.person_input.addItem(name)
        self.project_input.clear()
        for _, name in get_project():
            self.project_input.addItem(name)

    def load_list(self):
        TIME_CELL = 0
        PERSON_CELL = 1
        PROJECT_CELL = 2
        SUB_PROJECT_CELL = 3
        KIND_CELL = 4
        AMOUNT_CELL = 5
        MEMO_CELL = 6
        ACTION_CELL = 7

        rows = filter_transfer(*self.filters)

        if self.filters != ("", "", ""):
            self.filters_label.setText(f"{len(rows)} 条结果, 过滤条件: {','.join([x for x in self.filters if x])}")
        else:
            self.filters_label.setText(f"{len(rows)} 条结果, 未过滤")

        self.transfer_table.setRowCount(len(rows))

        for row, (id_, time, person, project, sub_project, kind, amount, memo) in enumerate(rows):
            action_cell_widget = QWidget()

            layout = QHBoxLayout(action_cell_widget)
            edit_btn = QPushButton('编辑')
            delete_btn = QPushButton("删除")
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(2)

            edit_btn.clicked.connect(partial(self.handle_edit, id_, time, person, project, sub_project, kind, amount, memo))
            delete_btn.clicked.connect(partial(self.handle_delete, id_))

            layout.addWidget(edit_btn)
            layout.addWidget(delete_btn)

            self.transfer_table.setItem(row, TIME_CELL, QTableWidgetItem(time))
            self.transfer_table.setItem(row, PERSON_CELL, QTableWidgetItem(person or ""))
            self.transfer_table.setItem(row, PROJECT_CELL, QTableWidgetItem(project or ""))
            self.transfer_table.setItem(row, SUB_PROJECT_CELL, QTableWidgetItem(sub_project or ""))
            self.transfer_table.setItem(row, KIND_CELL, QTableWidgetItem(kind or ""))
            self.transfer_table.setItem(row, AMOUNT_CELL, QTableWidgetItem(str(amount)))
            self.transfer_table.setItem(row, MEMO_CELL, QTableWidgetItem(memo or ""))
            self.transfer_table.setCellWidget(row, ACTION_CELL, action_cell_widget)

            self.transfer_table.setEditTriggers(QTableWidget.NoEditTriggers)

    def load(self):
        self.load_combo()
        self.load_list()


    def handle_add(self):
        time = self.time_input.dateTime().toString("yyyy-MM-dd HH:mm:ss")
        person = self.person_input.currentText()
        project = self.project_input.currentText()
        sub_project = self.sub_project_input.currentText()
        kind = self.kind_input.currentText()
        amount = self.amount_input.text()
        memo = self.memo_input.text()
        try:
            add_transfer(time, person, project, sub_project, kind, amount, memo)
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))
            return
        self.load_balance()
        self.load_list()

    def handle_edit(self, id_, time, person, project, sub_project, kind, amount, memo):
        dialog = EditTranferDialog(id_, time, person, project, sub_project, kind, amount, memo)
        if dialog.exec() == QDialog.Accepted:
            self.load_balance()
            self.load_list()

    def handle_delete(self, id_):
        reply = QMessageBox.question(self, "确认", f"确定要删除交易记录吗？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return

        try:
            delete_transfer(id_)
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))
            return
        self.load_balance()
        self.load_list()

    def export_to_excel(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "导出为 Excel 文件", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

        if not file_name:
            print('file name not specified')
            return

        wb = excel_from_table(self.transfer_table, "流水记录", False)
        msg_content = "导出成功，文件已保存到 {}".format(file_name)
        try:
            # Save the workbook to a file
            wb.save(file_name)
        except Exception as e:
            msg_content = "导出失败，错误信息：{}".format(str(e))

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(msg_content)
        msg.setWindowTitle("导出结果")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()

class SummaryTab(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout()
        self.btn = QPushButton("导出为 Excel")
        self.btn.clicked.connect(self.export_to_excel)
        self.summary_table = QTableWidget()
        self.summary_table.setEditTriggers(QTableWidget.NoEditTriggers)
        # self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.btn)
        layout.addWidget(self.summary_table)
        self.setLayout(layout)

    def get_project_order(self):
        project_list = [name for _, name in get_project()]
        project_map = defaultdict(list)
        for sub_project, project in get_sub_project():
            project_map[project].append(sub_project)
        r = []
        for parent in project_list:
            for sub_project in project_map[parent]:
                r.append(f'{parent}\n{sub_project}')
        return r


    def load(self):
        transfer_list = get_transfer()

        person_order = [name for _, name in get_person()]
        person_summary = defaultdict(lambda: defaultdict(decimal.Decimal))
        for person in person_order:
            person_summary[person]['入'] = 0
            person_summary[person]['出'] = 0

        project_order = self.get_project_order()
        project_summary = {project: 0 for project in project_order}

        summary = defaultdict(decimal.Decimal)
        for person in person_order:
            for project in project_order:
                summary[(person, '入', project)] = 0
                summary[(person, '出', project)] = 0

        for row, (id_, time, person, project, sub_project, kind, amount, memo) in enumerate(transfer_list):
            amount = decimal.Decimal(amount)
            summary[(person, kind, f'{project}\n{sub_project}')] += amount
            person_summary[person][kind] += amount
            project_summary[f'{project}\n{sub_project}'] += kind_sign(kind) * amount

        self.summary_table.clear()
        self.summary_table.clearSpans()

        horizontal_headers = project_order + ['人员统计', '人员合计']
        vertial_headers = []
        for person in person_order:
            vertial_headers.append(f'{person} 入')
            vertial_headers.append(f'出')
        vertial_headers.append('项目合计')

        self.summary_table.setRowCount(len(vertial_headers))
        self.summary_table.setColumnCount(len(horizontal_headers))
        self.summary_table.setVerticalHeaderLabels(vertial_headers)
        self.summary_table.setHorizontalHeaderLabels(horizontal_headers)            

        for col in range(self.summary_table.rowCount()):
            self.summary_table.verticalHeaderItem(col).setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)


        for col, project in enumerate(project_order):
            for x, person in enumerate(person_order):
                self.summary_table.setItem(x * 2, col, QTableWidgetItem(str(summary[(person, "入账", project)])))
                self.summary_table.setItem(x * 2 + 1, col, QTableWidgetItem(str(summary[(person, "出账", project)])))

        for x, person in enumerate(person_order):
            income = person_summary[person]["入账"]
            spend = person_summary[person]["出账"]
            self.summary_table.setItem(x * 2, len(project_summary), QTableWidgetItem(str(income)))
            self.summary_table.setItem(x * 2 + 1, len(project_summary), QTableWidgetItem(str(spend)))
            self.summary_table.setItem(x * 2, len(project_summary) + 1, QTableWidgetItem(str(income - spend)))
            self.summary_table.setSpan(x * 2, len(project_summary) + 1, 2, 1)
            print('set span', x * 2, len(project_summary) + 1, 2, 1)

        for col, project in enumerate(project_order):
            self.summary_table.setItem(2 * len(person_summary), col, QTableWidgetItem(str(project_summary[project])))

        self.summary_table.setItem(len(vertial_headers) - 1, len(project_summary), QTableWidgetItem(str(sum(project_summary.values()))))
        self.summary_table.setSpan(len(vertial_headers) - 1, len(project_summary), 1, 2)
        print('set span', len(vertial_headers) - 1, len(project_summary), 1, 2)

    def export_to_excel(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "导出为 Excel 文件", "", "Excel Files (*.xlsx);;All Files (*)", options=options)

        if not file_name:
            print('file name not specified')
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "流水统计"

        # Get the headers from the table and add them to the first and second row in the sheet
        for col in range(self.summary_table.columnCount()):
            header_lines = self.summary_table.horizontalHeaderItem(col).text().split('\n')
            ws.cell(row=1, column=col+3, value=header_lines[0])
            if len(header_lines) > 1:
                ws.cell(row=2, column=col+3, value=header_lines[1])

        # merge adjacent cells with identical value at first row
        project_start_col = 3
        for col in range(4, self.summary_table.columnCount() + 2):
            if ws.cell(row=1, column=col).value != ws.cell(row=1, column=col-1).value:
                ws.merge_cells(start_row=1, start_column=project_start_col, end_row=1, end_column=col - 1)
                ws.cell(row=1, column=project_start_col).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                project_start_col = col

        # Get the vertical headers (row headers) and add them to the first and second column
        for row in range(self.summary_table.rowCount()):
            header_sections = self.summary_table.verticalHeaderItem(row).text().split(' ', 1)
            if len(header_sections) == 2:
                ws.cell(row=row+3, column=1, value=header_sections[0])
                ws.cell(row=row+3, column=2, value=header_sections[1])
            else:
                ws.cell(row=row+3, column=2, value=header_sections[0])

        # merge adjacent cells that belong to same person
        for row in range(3, self.summary_table.rowCount() + 2, 2):
            ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)
            ws.cell(row=row, column=1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            ws.merge_cells(start_row=row, start_column=self.summary_table.columnCount() + 2, end_row=row+1, end_column=self.summary_table.columnCount() + 2)
            ws.cell(row=row, column=self.summary_table.columnCount() + 2).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # Get the table data and add it to the sheet starting from row 2
        for row in range(self.summary_table.rowCount()):
            for col in range(self.summary_table.columnCount()):
                item = self.summary_table.item(row, col)
                if item:
                    ws.cell(row=row+3, column=col+3, value= decimal.Decimal(item.text()))
                    ws.cell(row=row+3, column=col+3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # merge blank cells at topleft
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)

        # merge '项目合计'
        ws.cell(row=self.summary_table.rowCount() + 2, column=1, value='项目合计')
        ws.merge_cells(start_row=self.summary_table.rowCount() + 2, start_column=1, end_row=self.summary_table.rowCount() + 2, end_column=2)
        ws.cell(row=self.summary_table.rowCount() + 2, column=1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # merge '人员统计'
        ws.merge_cells(start_row=1, start_column=self.summary_table.columnCount() + 1, end_row=2, end_column=self.summary_table.columnCount() + 1)
        ws.cell(row=1, column=self.summary_table.columnCount() + 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # merge '人员合计'
        ws.merge_cells(start_row=1, start_column=self.summary_table.columnCount() + 2, end_row=2, end_column=self.summary_table.columnCount() + 2)
        ws.cell(row=1, column=self.summary_table.columnCount() + 2).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # merge bottomright 2 cells
        ws.merge_cells(start_row=self.summary_table.rowCount() + 2, start_column=self.summary_table.columnCount() + 1, end_row=self.summary_table.rowCount() + 2, end_column=self.summary_table.columnCount() + 2)
        ws.cell(row=self.summary_table.rowCount() + 2, column=self.summary_table.columnCount() + 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        msg_content = "导出成功，文件已保存到 {}".format(file_name)
        try:
            # Save the workbook to a file
            wb.save(file_name)
        except Exception as e:
            msg_content = "导出失败，错误信息：{}".format(str(e))

        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(msg_content)
        msg.setWindowTitle("导出结果")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()

class SettingTab(QWidget):
    def __init__(self):
        super().__init__()
        form = QFormLayout()
        font_combo = QComboBox()
        font_combo.addItems(['小', '中', '大'])
        font_combo.currentTextChanged.connect(self.change_font)
        form.addRow("字体大小:", font_combo)
        self.setLayout(form)


    def change_font(self, size):
        font = QFont()
        if size == "小":
            font.setPointSize(10)
        elif size == "中":
            font.setPointSize(14)
        elif size == "大":
            font.setPointSize(18)
        QApplication.setFont(font)

    def load(self):
        pass


class LedgerApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setMinimumSize(1000, 800)
        self.setWindowTitle("账本应用")

        layout = QVBoxLayout()

        self.tabs = QTabWidget()
        self.tabs.addTab(PersonTab(), "人员")
        self.tabs.addTab(ProjectTab(), "项目")
        self.tabs.addTab(TransferTab(), "流水")
        self.tabs.addTab(SummaryTab(), "统计")
        self.tabs.addTab(SettingTab(), "设置")

        self.tabs.currentChanged.connect(self.on_tab_changed)

        layout.addWidget(self.tabs)
        self.setLayout(layout)

    def on_tab_changed(self, index):
        print('tab change', self.tabs.tabText(index))
        self.tabs.widget(index).load()

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--drop":
        init_db(True)
    else:
        init_db(False)

    app = QApplication(sys.argv)
    window = LedgerApp()
    window.show()
    sys.exit(app.exec())
